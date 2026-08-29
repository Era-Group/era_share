# -*- coding: utf-8 -*-

import base64
from io import BytesIO

from odoo import models, fields, api ,_
from odoo.exceptions import UserError
from odoo.tools import float_compare, float_is_zero


class UmrahVisa(models.Model):
    _name = 'umrah.visa'
    _description = 'Umrah Visa'
    _rec_name = 'visa_number'
    _inherit = ['mail.thread','mail.activity.mixin']

    # Basic Information
    visa_number = fields.Char(string='Visa Number')
    mofa_number = fields.Char(string='MOFA Number', copy=False)
    visa_type = fields.Selection([
        ('umrah', 'Umrah Visa'),
        ('hajj', 'Hajj Visa'),
        ('visit', 'Visit Visa'),
        ('business', 'Business Visa')
    ], string='Visa Type', required=True, default='umrah')
    
    # Applicant Information
    pilgrim_id = fields.Many2one('umrah.pilgrim', string='Pilgrim', required=True)
    
    # Application Details
    application_date = fields.Date(string='Application Date', required=True, default=fields.Date.today)
    application_reference = fields.Char(string='Application Reference')
    consulate_location = fields.Char(string='Consulate/Embassy Location')
    
    # Visa Details
    issue_date = fields.Date(string='Issue Date')
    expiry_date = fields.Date(string='Expiry Date')
    validity_period = fields.Integer(string='Validity Period (Days)', compute='_compute_validity_period', store=True)
    
    entry_type = fields.Selection([
        ('single', 'Single Entry'),
        ('multiple', 'Multiple Entry')
    ], string='Entry Type', default='single')
    
    duration_of_stay = fields.Integer(string='Duration of Stay (Days)', default=30)
    
    # Status
    status = fields.Selection([
        ('draft', 'Draft'),
        ('submitted', 'Submitted'),
        ('under_review', 'Under Review'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected'),
        ('issued', 'Issued'),
        ('invoiced', 'Invoiced'),
        ('expired', 'Expired'),
        ('cancelled', 'Cancelled')
    ], string='Status', default='draft', tracking=True)
    
    # Fees
    visa_fee = fields.Float(string='Visa Fee')
    service_fee = fields.Float(string='Service Fee')
    total_fee = fields.Float(string='Total Fee', compute='_compute_total_fee', store=True)
    
    # Payment Information
    payment_status = fields.Selection([
        ('unpaid', 'Unpaid'),
        ('partial', 'Partially Paid'),
        ('paid', 'Paid'),
        ('refunded', 'Refunded')
    ], string='Payment Status', default='unpaid')
    
    payment_date = fields.Date(string='Payment Date')
    payment_reference = fields.Char(string='Payment Reference')
    
    # Documents
    passport_copy_submitted = fields.Boolean(string='Passport Copy Submitted')
    photo_submitted = fields.Boolean(string='Photo Submitted')
    application_form_submitted = fields.Boolean(string='Application Form Submitted')
    invitation_letter_submitted = fields.Boolean(string='Invitation Letter Submitted')
    
    # Relations
    trip_id = fields.Many2one('umrah.trip', string='Trip')
    group_id = fields.Many2one('umrah.group', string='Group')
    agent_id = fields.Many2one('umrah.agent', string='Processing Agent')
    
    # Additional Information
    rejection_reason = fields.Text(string='Rejection Reason')
    notes = fields.Text(string='Notes')
    created_date = fields.Datetime(string='Created Date', default=fields.Datetime.now)

    currency_id = fields.Many2one(
        'res.currency',
        string='Currency',
        default=lambda self: self.env.company.currency_id,
    )

    # =====================================================
    # Section 5 - Billing fields
    # =====================================================

    invoice_id = fields.Many2one('account.move', string='Invoice', readonly=True)
    invoice_number = fields.Char(
        related='invoice_id.name', string='Invoice Number', store=True
    )

    agent_number = fields.Char(
        related='agent_id.agent_number', string='Agent Number', store=True
    )
    sub_agent_id = fields.Many2one('umrah.agent', string='Sub-Agent')

    # Reflects which invoice branch will be used (display only).
    agent_billing_type = fields.Selection(
        related='agent_id.agent_billing_type',
        string='Agent Billing Type',
        readonly=True,
    )

    visa_count = fields.Integer(string='Visa Count')

    # FROZEN prices: snapshotted from the agent's active pricing line at
    # create/issue time. STORED & writable - NOT a live related to the
    # agent's current_* fields, so an existing visa never re-prices when
    # the agent's pricing lines change later.
    sale_price = fields.Monetary(
        string='Sale Price', currency_field='currency_id'
    )
    purchase_price = fields.Monetary(
        string='Purchase Price', currency_field='currency_id'
    )

    total = fields.Monetary(
        string='Total', compute='_compute_total', store=True,
        currency_field='currency_id',
    )

    nusuk_file = fields.Binary(string='Nusuk Visa File')
    nusuk_filename = fields.Char(string='Nusuk File Name')
    nusuk_total = fields.Monetary(
        string='Nusuk Visa Fees', currency_field='currency_id'
    )
    nusuk_grand_total = fields.Monetary(
        string='Nusuk Grand Total', currency_field='currency_id'
    )

    # ---------------------------------------------------------------------
    # Nusuk column indices (1-based).
    #
    # CONFIRMED 2026-07-20: the computed total (sale_price x visa_count) is
    # checked against BOTH columns — visa fees (col 14, رسوم التأشيرة) and
    # the grand total (col 24, المبلغ الاجمالي = ground + insurance +
    # transport + visa). The check passes when either column matches, since
    # which one applies depends on what the agent's sale price covers; the
    # mismatch warning always reports both parsed values.
    # Reference sums (Voucher-List_1777895344591__1_.xlsx, 84 data rows):
    # col 14 = 206100.00, col 24 = 330501.63.
    # ---------------------------------------------------------------------
    # Defaults; overridable from Settings (era_nusuk.nusuk_col_* params).
    _NUSUK_COL_VISA_FEES = 14    # رسوم التأشيرة — visa fees only
    _NUSUK_COL_GRAND_TOTAL = 24  # المبلغ الاجمالي — all services

    def _get_nusuk_columns(self):
        """Configurable 1-based column indices for the Nusuk file."""
        icp = self.env["ir.config_parameter"].sudo()
        fees = int(icp.get_param(
            "era_nusuk.nusuk_col_visa_fees", self._NUSUK_COL_VISA_FEES))
        grand = int(icp.get_param(
            "era_nusuk.nusuk_col_grand_total", self._NUSUK_COL_GRAND_TOTAL))
        return (fees, grand)

    @api.depends('sale_price', 'visa_count')
    def _compute_total(self):
        for rec in self:
            rec.total = (rec.sale_price or 0.0) * (rec.visa_count or 0)

    # --------------------------------------------------
    # Price freezing
    # --------------------------------------------------

    def _get_price_snapshot(self, date=None):
        """Return (sale_price, purchase_price) from the agent's active line."""
        self.ensure_one()
        if not self.agent_id:
            return (0.0, 0.0)
        line = self.agent_id._get_active_price(date)
        if not line:
            return (0.0, 0.0)
        return (line.sale_price, line.purchase_price)

    @api.onchange('agent_id')
    def _onchange_agent_snapshot_prices(self):
        # Refresh frozen prices while still editable (draft only).
        if self.status == 'draft':
            ref_date = self.issue_date or self.application_date
            self.sale_price, self.purchase_price = self._get_price_snapshot(ref_date)

    # --------------------------------------------------
    # Nusuk file parsing (Section 5.2)
    # --------------------------------------------------

    def _normalize_number(self, raw):
        """Normalize a Nusuk numeric cell (Arabic decimal/thousands, bidi
        marks, stray spaces) into a float. Reused for every numeric column."""
        if raw is None:
            return 0.0
        if isinstance(raw, (int, float)):
            return float(raw)
        s = str(raw)
        # strip bidi / RTL marks
        for mark in ('‏', '‎', '‪', '‫', '‬'):
            s = s.replace(mark, '')
        s = s.replace('٫', '.')   # Arabic decimal separator -> dot
        s = s.replace('٬', '')    # Arabic thousands separator -> drop
        s = s.replace(',', '')          # Latin thousands separator -> drop
        s = s.replace('\xa0', '').replace(' ', '').strip()
        if not s:
            return 0.0
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    def _parse_nusuk_totals(self):
        """Sum both the visa-fees column (col 14, رسوم التأشيرة) and the grand
        total column (col 24, المبلغ الاجمالي) across all data rows (header in
        row 1, data from row 2, no totals row). Returns (visa_fees, grand)."""
        self.ensure_one()
        if not self.nusuk_file:
            return (0.0, 0.0)
        from openpyxl import load_workbook
        data = base64.b64decode(self.nusuk_file)
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        try:
            ws = wb.active
            fees_col, grand_col = self._get_nusuk_columns()
            fees = grand = 0.0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None or all(c in (None, '') for c in row):
                    continue
                if len(row) >= fees_col:
                    fees += self._normalize_number(row[fees_col - 1])
                if len(row) >= grand_col:
                    grand += self._normalize_number(row[grand_col - 1])
            return (fees, grand)
        finally:
            wb.close()

    def _parse_nusuk_total(self):
        """Visa-fees sum only (kept for backwards compatibility)."""
        return self._parse_nusuk_totals()[0]

    @api.onchange('nusuk_file')
    def _onchange_check_nusuk_total(self):
        if not self.nusuk_file:
            self.nusuk_total = 0.0
            self.nusuk_grand_total = 0.0
            return
        try:
            self.nusuk_total, self.nusuk_grand_total = self._parse_nusuk_totals()
        except Exception as exc:  # noqa: BLE001 - surface any parse error to user
            self.nusuk_total = 0.0
            self.nusuk_grand_total = 0.0
            return {
                'warning': {
                    'title': _('Nusuk File Error'),
                    'message': _(
                        'Could not parse the uploaded Nusuk file: %s'
                    ) % exc,
                }
            }
        if not self.nusuk_total and not self.nusuk_grand_total:
            # Empty / unparseable file -> nothing to compare against.
            return

        # Case (a): no computed total means no sale price is configured for the
        # agent (the frozen snapshot is 0). This is a config issue, NOT a data
        # mismatch - tell the user precisely that, so they don't chase a
        # non-existent discrepancy.
        if float_is_zero(self.total, precision_digits=2):
            return {
                'warning': {
                    'title': _('Cannot Verify Total / تعذّر التحقق من الإجمالي'),
                    'message': _(
                        'Cannot verify — no sale price is set for this agent. '
                        'Add an active price line on the agent\'s Pricing tab '
                        'first.\n\n'
                        'تعذّر التحقق — لا يوجد سعر بيع محدد لهذا الوكيل. '
                        'يرجى إضافة سطر سعر فعّال في تبويب التسعير الخاص '
                        'بالوكيل أولاً.'
                    ),
                }
            }

        # Case (b): mismatch only when the computed total matches NEITHER the
        # visa-fees column nor the grand-total column (confirmed 2026-07-20:
        # check against both, warn with both values).
        matches_fees = float_compare(
            self.nusuk_total, self.total, precision_digits=2
        ) == 0
        matches_grand = float_compare(
            self.nusuk_grand_total, self.total, precision_digits=2
        ) == 0
        if not matches_fees and not matches_grand:
            return {
                'warning': {
                    'title': _('Total Mismatch / اختلاف في الإجمالي'),
                    'message': _(
                        'The computed total (%(computed)s) matches neither the '
                        'visa-fees total (%(fees)s) nor the grand total '
                        '(%(grand)s) in the Nusuk file.\n\n'
                        'الإجمالي المحتسب (%(computed)s) لا يطابق إجمالي رسوم '
                        'التأشيرة (%(fees)s) ولا المبلغ الإجمالي (%(grand)s) '
                        'في ملف نسك.'
                    ) % {
                        'computed': self.total,
                        'fees': self.nusuk_total,
                        'grand': self.nusuk_grand_total,
                    },
                }
            }

    # --------------------------------------------------
    # Invoice generation (Section 5.3)
    # --------------------------------------------------

    def _get_visa_sale_tax(self):
        """Return the sale VAT for the service-fee line. The percentage is
        configurable (Settings > Era Nusuk, default 15%).

        Hard-fails if it is not configured: in the Saudi/ZATCA context a
        silently untaxed invoice is a compliance risk, so we refuse to post
        rather than omit VAT.
        """
        self.ensure_one()
        percent = float(self.env["ir.config_parameter"].sudo().get_param(
            "era_nusuk.service_vat_percent", 15.0))
        tax = self.env['account.tax'].search([
            ('type_tax_use', '=', 'sale'),
            ('amount', '=', percent),
            ('company_id', '=', self.env.company.id),
        ], limit=1)
        if not tax:
            raise UserError(
                _("VAT %s%% tax not configured for this company") % percent)
        return tax

    # =====================================================================
    # CONFIRMED 2026-07-20: for a 'default' agent the invoice is TWO lines
    # totalling the sale price:
    #   1. Visa fees (government, recharged at purchase price) — no VAT,
    #      it is a disbursement recovered on the agent's behalf.
    #   2. Service fee (sale - purchase margin) — 15% VAT, this is the
    #      company's own taxable revenue.
    # The old literal-spec third "Sale Price" line double-counted the total
    # (2 x sale price) and was removed.
    # =====================================================================
    def action_create_visa_invoice(self):
        self.ensure_one()
        if not self.agent_id:
            raise UserError(_("Set an Agent before creating the invoice."))
        if not self.agent_id.partner_id:
            raise UserError(_("The selected Agent has no linked partner."))

        partner = self.agent_id.partner_id
        qty = self.visa_count or 0
        billing = self.agent_id.agent_billing_type

        if billing == 'actual':
            # Actual agent pays the visa fees only - single line, no tax.
            lines = [(0, 0, {
                'name': _('Visa Fees'),
                'quantity': qty,
                'price_unit': self.purchase_price,
            })]
        else:
            # Default agent: visa fees (no VAT) + service margin (15% VAT).
            diff = (self.sale_price or 0.0) - (self.purchase_price or 0.0)
            tax = self._get_visa_sale_tax()
            lines = [
                (0, 0, {
                    'name': _('Visa Fees'),
                    'quantity': qty,
                    'price_unit': self.purchase_price,
                }),
                (0, 0, {
                    'name': _('Service Fee'),
                    'quantity': qty,
                    'price_unit': diff,
                    'tax_ids': [(6, 0, tax.ids)],
                }),
            ]

        move = self.env['account.move'].create({
            'move_type': 'out_invoice',
            'partner_id': partner.id,
            'invoice_date': self.issue_date,
            'umrah_visa_id': self.id,
            'agent_id': self.agent_id.id,
            'invoice_line_ids': lines,
        })
        self.invoice_id = move.id
        self.status = 'invoiced'

        return {
            'name': _('Invoice'),
            'type': 'ir.actions.act_window',
            'res_model': 'account.move',
            'view_mode': 'form',
            'res_id': move.id,
        }

    # deprecated: replaced by action_create_visa_invoice in Section 5.1/5.3
    def action_create_agent_invoice(self):
        self.ensure_one()
        if not self.agent_id:
            raise UserError(_("يجب تحديد وكيل لإنشاء الفاتورة"))
        invoice_lines = []
        invoice_lines.append((0, 0, {
            'name': f"{self.pilgrim_id.full_name} - {self.visa_number}",
            'quantity': 1,
            'price_unit': self.total_fee,
        }))

        invoice = self.env['account.move'].create({
            'move_type': 'out_invoice',
            'partner_id': self.agent_id.partner_id.id,
            'invoice_line_ids': invoice_lines,
        })

        return {
            'name': 'الفاتورة',
            'type': 'ir.actions.act_window',
            'res_model': 'account.move',
            'view_mode': 'form',
            'res_id': invoice.id,
        }
    
    @api.depends('issue_date', 'expiry_date')
    def _compute_validity_period(self):
        for record in self:
            if record.issue_date and record.expiry_date:
                delta = record.expiry_date - record.issue_date
                record.validity_period = delta.days
            else:
                record.validity_period = 0
    
    @api.depends('visa_fee', 'service_fee')
    def _compute_total_fee(self):
        for record in self:
            record.total_fee = record.visa_fee + record.service_fee
    
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get('visa_number'):
                vals['visa_number'] = self.env['ir.sequence'].next_by_code('umrah.visa')
        records = super().create(vals_list)
        # Initial price snapshot for early display; the authoritative freeze
        # happens on issue (action_issue).
        for rec in records:
            if rec.agent_id and not rec.sale_price and not rec.purchase_price:
                ref_date = rec.issue_date or rec.application_date
                rec.sale_price, rec.purchase_price = rec._get_price_snapshot(
                    ref_date
                )
        return records
    
    # ------------------------------------------------------------------
    # Regulatory pre-submission checks (Nusuk Masar rules). Each check is
    # toggleable from Settings so the flow stays flexible.
    # ------------------------------------------------------------------

    def _check_masar_requirements(self):
        icp = self.env["ir.config_parameter"].sudo()
        enforce_hotel = icp.get_param(
            "era_nusuk.enforce_hotel_coverage", "True") == "True"
        enforce_transport = icp.get_param(
            "era_nusuk.enforce_transport_before_visa", "True") == "True"
        enforce_quota = icp.get_param(
            "era_nusuk.enforce_agent_quota", "False") == "True"
        for rec in self:
            group = rec.group_id
            if enforce_hotel and group:
                if not group._hotel_coverage_ok():
                    raise UserError(_(
                        "Masar rule: group %s has no approved hotel bookings "
                        "covering the full stay (arrival to departure). "
                        "Complete the hotel bookings first, or disable this "
                        "check in Settings.") % group.display_name)
            if enforce_transport and group:
                movements = self.env["umrah.movement"].search([
                    ("group_ids", "in", group.id)])
                types = set(movements.mapped("movement_type"))
                if not {"arrival", "departure"} <= types:
                    raise UserError(_(
                        "Masar rule: group %s must have both an arrival and "
                        "a departure movement registered before visa "
                        "submission, or disable this check in Settings.")
                        % group.display_name)
            if enforce_quota and rec.agent_id and rec.agent_id.seasonal_quota:
                rec.agent_id.invalidate_recordset(
                    ['quota_used', 'quota_remaining'])
                if rec.agent_id.quota_remaining <= 0:
                    raise UserError(_(
                        "Agent %(agent)s has exhausted his seasonal quota "
                        "(%(quota)s visas).") % {
                            "agent": rec.agent_id.display_name,
                            "quota": rec.agent_id.seasonal_quota})

    def action_submit(self):
        self._check_masar_requirements()
        self.status = 'submitted'
    
    def action_approve(self):
        self.status = 'approved'
    
    def action_reject(self):
        self.status = 'rejected'
    
    def action_issue(self):
        for rec in self:
            rec.status = 'issued'
            if not rec.issue_date:
                rec.issue_date = fields.Date.today()
            # Authoritative price freeze: lock in the agent's active price as
            # of the issue date. Stored writable fields -> never re-priced later.
            rec.sale_price, rec.purchase_price = rec._get_price_snapshot(
                rec.issue_date
            )

    def action_cancel(self):
        self.status = 'cancelled'







    def _create_agent_invoices(self):
        """Create one customer invoice per agent for the visas in ``self``,
        with one line per visa (priced at ``total_fee``). Marks each visa
        invoiced and links it to its invoice. Returns an act_window on the
        created invoices."""
        invoices = self.env['account.move']
        for agent, visas in self.grouped('agent_id').items():
            if not agent or not agent.partner_id:
                continue
            invoice = self.env['account.move'].create({
                'move_type': 'out_invoice',
                'partner_id': agent.partner_id.id,
                'agent_id': agent.id,
                'invoice_line_ids': [(0, 0, {
                    'name': f"{visa.pilgrim_id.full_name} - {visa.visa_number}",
                    'quantity': 1,
                    'price_unit': visa.total_fee,
                }) for visa in visas],
            })
            visas.write({'status': 'invoiced', 'invoice_id': invoice.id})
            invoices += invoice

        if not invoices:
            raise UserError(_("None of the selected visas is linked to an agent."))

        return {
            'name': _('Invoices'),
            'type': 'ir.actions.act_window',
            'res_model': 'account.move',
            'view_mode': 'list,form',
            'domain': [('id', 'in', invoices.ids)],
        }

    def action_create_bulk_agent_invoices(self):
        """Legacy list-selection entry point (kept for server actions)."""
        selected_visas = self.env['umrah.visa'].browse(
            self.env.context.get('active_ids', [])) or self
        selected_visas = selected_visas.filtered(
            lambda v: v.status not in ('cancelled', 'rejected', 'invoiced'))
        if not selected_visas:
            raise UserError(_("No visas selected to invoice."))
        return selected_visas._create_agent_invoices()
